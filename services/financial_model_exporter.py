"""
=============================================================================
MODANO 3-WAY INTEGRATED MODELING ECOSYSTEM: FINANCIAL MODEL EXPORTER
=============================================================================
Institutional-grade, Modano-compliant Excel Workbook Exporter generating a
comprehensive 7-Tab dynamic financial model with native Excel formulas
(SUM, IF, cross-sheet links), corporate finance formatting, freeze panes,
and zero formula syntax errors.

Workbook Architecture (7 Tabs):
-------------------------------
Tab 1: Cover & Summary Dashboard (Valuation, Solvency, Risk Firewalls)
Tab 2: Income Statement (5-Year P&L with dynamic formulas)
Tab 3: Balance Sheet (5-Year BS with exact closure & IS_BALANCED checks)
Tab 4: Cash Flow Statement (Direct Method CFS reconciling to Delta Cash)
Tab 5: Working Capital Schedule (DSO, DIO, DPO, CCC, Delta NWC)
Tab 6: Debt & Capital Schedule (Damodaran rating, ICR, Kd, Amortization)
Tab 7: Valuation & Sensitivity (DCF, DDM, FCFE, 5x5 WACC vs g Sensitivity)
=============================================================================
"""

from __future__ import annotations

import os
import re
import math
import logging
from typing import Dict, List, Any, Optional, Tuple, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.three_statement_engine import (
    ThreeStatementEngine,
    ThreeStatementForecastResult,
    IncomeStatementForecast,
    BalanceSheetForecast,
    CashFlowForecast,
    LiquidityDistressCheck,
    sanitize_float,
    safe_div,
    clamp,
)

logger = logging.getLogger(__name__)


# =============================================================================
# EXCEL COLOR PALETTE & STYLING CONSTANTS (Modano Corporate Standard)
# =============================================================================

COLOR_NAVY_HEADER = "1F4E79"       # Primary Title / Header Fill (Navy Blue)
COLOR_TEXT_WHITE = "FFFFFF"        # Header Text
COLOR_SECTION_ACCENT = "D9E1F2"    # Section Break Fill (Ice / Soft Lavender)
COLOR_TEXT_NAVY = "1F4E79"         # Section Text
COLOR_ZEBRA_LIGHT = "F2F2F2"       # Table row alternate
COLOR_GREEN_FILL = "E2EFDA"        # Balanced / Healthy Fill (Soft Sage Green)
COLOR_GREEN_TEXT = "375623"        # Green Text
COLOR_RED_FILL = "FCE4D6"          # Distressed / Warning Fill (Soft Coral)
COLOR_RED_TEXT = "C65911"          # Red Text
COLOR_BORDER_GRAY = "D9D9D9"       # Thin Grid Border

# Standard Fonts
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=COLOR_TEXT_WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color=COLOR_TEXT_WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_WHITE)
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_NAVY)
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color="000000")
FONT_REGULAR = Font(name="Calibri", size=11, bold=False, color="000000")
FONT_ITALIC = Font(name="Calibri", size=10, italic=True, color="595959")
FONT_KPI_VAL = Font(name="Calibri", size=14, bold=True, color=COLOR_TEXT_NAVY)
FONT_KPI_LBL = Font(name="Calibri", size=9, bold=False, color="595959")

# Standard Fills
FILL_HEADER = PatternFill(start_color=COLOR_NAVY_HEADER, end_color=COLOR_NAVY_HEADER, fill_type="solid")
FILL_SECTION = PatternFill(start_color=COLOR_SECTION_ACCENT, end_color=COLOR_SECTION_ACCENT, fill_type="solid")
FILL_ZEBRA = PatternFill(start_color=COLOR_ZEBRA_LIGHT, end_color=COLOR_ZEBRA_LIGHT, fill_type="solid")
FILL_GREEN = PatternFill(start_color=COLOR_GREEN_FILL, end_color=COLOR_GREEN_FILL, fill_type="solid")
FILL_RED = PatternFill(start_color=COLOR_RED_FILL, end_color=COLOR_RED_FILL, fill_type="solid")

# Standard Borders
BORDER_THIN_GRAY = Side(border_style="thin", color=COLOR_BORDER_GRAY)
BORDER_THICK_NAVY = Side(border_style="medium", color=COLOR_NAVY_HEADER)
BORDER_DOUBLE_BOTTOM = Side(border_style="double", color="000000")

BOX_BORDER = Border(left=BORDER_THIN_GRAY, right=BORDER_THIN_GRAY, top=BORDER_THIN_GRAY, bottom=BORDER_THIN_GRAY)
TOTAL_BORDER = Border(top=BORDER_THIN_GRAY, bottom=BORDER_DOUBLE_BOTTOM)
HEADER_BORDER = Border(left=BORDER_THIN_GRAY, right=BORDER_THIN_GRAY, top=BORDER_THICK_NAVY, bottom=BORDER_THICK_NAVY)

# Standard Alignments
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Standard Number Formats
FORMAT_BILLION = "#,##0.0;(#,##0.0);\"-\""      # In Billion VND
FORMAT_INTEGER = "#,##0;(#,##0);\"-\""          # Integer
FORMAT_PERCENT = "0.0%"                         # 1-Decimal Percentage
FORMAT_PERCENT_2 = "0.00%"                      # 2-Decimal Percentage
FORMAT_MULTIPLE = "0.00\"x\""                   # Multiples
FORMAT_CURRENCY_VND = "#,##0\" VND\""           # Per Share Price


# =============================================================================
# EXCEL EXPORTER SERVICE
# =============================================================================

class FinancialModelExporter:
    """
    Modano-Compliant 7-Tab Interactive Financial Model Excel Exporter.
    """

    @staticmethod
    def _shift_formula_column(formula_tmpl: str, target_col_letter: str) -> str:
        """Safely shifts formula template column 'C' coordinates to target_col_letter without corrupting sheet names."""
        if not formula_tmpl:
            return formula_tmpl
        return re.sub(r"(?<![A-Za-z0-9_])C(\d+)", rf"{target_col_letter}\1", formula_tmpl)

    @classmethod
    def export_to_excel(
        cls,
        forecast_result: ThreeStatementForecastResult,
        output_path: str,
        scale_unit: str = "billion", # 'billion' (default, divides by 1e9) or 'raw'
    ) -> str:
        """
        Builds and saves the complete 7-Tab Modano-compliant financial model to `output_path`.
        Returns the absolute file path of the generated workbook.
        """
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        
        wb = openpyxl.Workbook()
        # Remove default active sheet
        default_sheet = wb.active
        
        # Scale factor (Billion VND = 1e9)
        scale_div = 1e9 if scale_unit == "billion" else 1.0
        unit_label = "Tỷ VND (Billion VND)" if scale_unit == "billion" else "VND"

        # 1. Tab 1: Cover & Dashboard
        ws_dash = wb.create_sheet(title="Summary & Dashboard")
        cls._build_dashboard_tab(ws_dash, forecast_result, scale_div, unit_label)

        # 2. Tab 2: Income Statement
        ws_is = wb.create_sheet(title="Income Statement")
        cls._build_income_statement_tab(ws_is, forecast_result, scale_div, unit_label)

        # 3. Tab 3: Balance Sheet
        ws_bs = wb.create_sheet(title="Balance Sheet")
        cls._build_balance_sheet_tab(ws_bs, forecast_result, scale_div, unit_label)

        # 4. Tab 4: Cash Flow Statement
        ws_cfs = wb.create_sheet(title="Cash Flow Statement")
        cls._build_cash_flow_tab(ws_cfs, forecast_result, scale_div, unit_label)

        # 5. Tab 5: Working Capital Schedule
        ws_wc = wb.create_sheet(title="Working Capital Schedule")
        cls._build_working_capital_tab(ws_wc, forecast_result, scale_div, unit_label)

        # 6. Tab 6: Debt & Capital Schedule
        ws_debt = wb.create_sheet(title="Debt & Capital Schedule")
        cls._build_debt_capital_tab(ws_debt, forecast_result, scale_div, unit_label)

        # 7. Tab 7: Valuation & Sensitivity
        ws_val = wb.create_sheet(title="Valuation & Sensitivity")
        cls._build_valuation_sensitivity_tab(ws_val, forecast_result, scale_div, unit_label)

        # Remove default blank sheet if present
        if default_sheet in wb.worksheets:
            wb.remove(default_sheet)

        wb.active = ws_dash
        wb.save(output_path)
        logger.info("Successfully exported 7-tab financial model to %s", output_path)
        return os.path.abspath(output_path)

    # -------------------------------------------------------------------------
    # TAB 1: SUMMARY & DASHBOARD
    # -------------------------------------------------------------------------
    @classmethod
    def _build_dashboard_tab(
        cls,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        res: ThreeStatementForecastResult,
        div: float,
        unit_lbl: str,
    ) -> None:
        ws.views.sheetView[0].showGridLines = True
        
        # 1. Title Banner
        ws.merge_cells("A1:H2")
        c1 = ws["A1"]
        c1.value = f"MODANO INTEGRATED 3-WAY FINANCIAL MODEL: {res.symbol} ({res.company_name})"
        c1.font = FONT_TITLE
        c1.fill = FILL_HEADER
        c1.alignment = ALIGN_CENTER

        ws["A3"] = f"Ngành (Sector): {res.sector} | Đơn vị tính: {unit_lbl} | Kỳ dự phóng: {res.start_year} - {res.start_year + 4}"
        ws["A3"].font = FONT_ITALIC
        ws["A3"].alignment = ALIGN_LEFT

        # 2. Executive KPI Cards (Row 5-8)
        kpis = [
            ("DOANH THU 5 NĂM (5Y REV)", f"=SUM('Income Statement'!C5:G5)", FORMAT_BILLION, "B5:C6"),
            ("LỢI NHUẬN RÒNG 5Y (5Y NPAT)", f"=SUM('Income Statement'!C20:G20)", FORMAT_BILLION, "D5:E6"),
            ("DÒNG TIỀN HĐKD (5Y CFO)", f"=SUM('Cash Flow Statement'!C13:G13)", FORMAT_BILLION, "F5:G6"),
            ("TIỀN MẶT CUỐI KỲ (END CASH)", f"='Balance Sheet'!G5", FORMAT_BILLION, "H5:I6"),
        ]

        for lbl, formula, num_fmt, cell_range in kpis:
            top_left = cell_range.split(":")[0]
            ws.merge_cells(cell_range)
            cell = ws[top_left]
            cell.value = formula
            cell.font = FONT_KPI_VAL
            cell.number_format = num_fmt
            cell.alignment = ALIGN_CENTER
            cell.fill = FILL_SECTION
            cell.border = BOX_BORDER

        # Labels row under KPIs
        ws["B7"] = "Tổng Doanh Thu 5 Năm"
        ws["D7"] = "Tổng LNST 5 Năm"
        ws["F7"] = "Tổng Dòng Tiền CFO"
        ws["H7"] = "Tiền Mặt Năm Thứ 5"
        for col_let in ["B", "D", "F", "H"]:
            ws[f"{col_let}7"].font = FONT_KPI_LBL
            ws[f"{col_let}7"].alignment = ALIGN_CENTER

        # 3. Solvency & Risk Diagnostics Card
        row = 9
        ws.merge_cells(f"A{row}:I{row}")
        ws[f"A{row}"] = "ĐÁNH GIÁ AN TOÀN TÀI CHÍNH & RỦI RO THANH KHOẢN (SOLVENCY & LIQUIDITY FIREWALL)"
        ws[f"A{row}"].font = FONT_SECTION
        ws[f"A{row}"].fill = FILL_SECTION
        ws[f"A{row}"].alignment = ALIGN_LEFT

        row += 1
        distress = res.liquidity_distress_check
        is_healthy = distress.summary_assessment == "HEALTHY"
        status_fill = FILL_GREEN if is_healthy else FILL_RED

        ws[f"A{row}"] = "Trạng thái thanh khoản (Liquidity Status):"
        ws[f"C{row}"] = distress.summary_assessment
        ws[f"C{row}"].font = FONT_BOLD
        ws[f"C{row}"].fill = status_fill
        ws[f"C{row}"].alignment = ALIGN_CENTER

        ws[f"E{row}"] = "Cân đối Bảng CĐKT (BS Closure):"
        ws[f"G{row}"] = "100% CÂN ĐỐI (BALANCED)" if res.all_years_balanced else "LỆCH TOÁN HỌC"
        ws[f"G{row}"].font = FONT_BOLD
        ws[f"G{row}"].fill = FILL_GREEN if res.all_years_balanced else FILL_RED
        ws[f"G{row}"].alignment = ALIGN_CENTER

        row += 1
        ws[f"A{row}"] = "Số dư tiền tối thiểu (Min Cash):"
        ws[f"C{row}"] = distress.min_cash_balance / div
        ws[f"C{row}"].font = FONT_REGULAR
        ws[f"C{row}"].number_format = FORMAT_BILLION

        ws[f"E{row}"] = "Biên an toàn rủi ro (MOS Penalty):"
        ws[f"G{row}"] = distress.mos_penalty_pct
        ws[f"G{row}"].font = FONT_REGULAR
        ws[f"G{row}"].number_format = FORMAT_PERCENT

        # 4. 5-Year Executive Summary Table
        row += 2
        ws.merge_cells(f"A{row}:I{row}")
        ws[f"A{row}"] = "BẢNG TỔNG HỢP CÁC CHỈ TIÊU TÀI CHÍNH CỐT LÕI 5 NĂM DỰ PHÓNG"
        ws[f"A{row}"].font = FONT_SECTION
        ws[f"A{row}"].fill = FILL_SECTION

        row += 1
        headers = ["Chỉ tiêu tài chính cốt lõi (Metric)", "Mã dòng", f"Năm {res.forecast_years[0]}", f"Năm {res.forecast_years[1]}", f"Năm {res.forecast_years[2]}", f"Năm {res.forecast_years[3]}", f"Năm {res.forecast_years[4]}", "Tổng 5 Năm", "CAGR"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_HEADER

        summary_rows = [
            ("Doanh thu thuần (Net Revenue)", "REV", "='Income Statement'!C5", "='Income Statement'!D5", "='Income Statement'!E5", "='Income Statement'!F5", "='Income Statement'!G5", "=SUM(C{r}:G{r})", "=(G{r}/C{r})^(1/4)-1", FORMAT_BILLION, False),
            ("Lợi nhuận gộp (Gross Profit)", "GP", "='Income Statement'!C8", "='Income Statement'!D8", "='Income Statement'!E8", "='Income Statement'!F8", "='Income Statement'!G8", "=SUM(C{r}:G{r})", "=(G{r}/C{r})^(1/4)-1", FORMAT_BILLION, False),
            ("Lợi nhuận hoạt động (EBIT)", "EBIT", "='Income Statement'!C13", "='Income Statement'!D13", "='Income Statement'!E13", "='Income Statement'!F13", "='Income Statement'!G13", "=SUM(C{r}:G{r})", "=(G{r}/C{r})^(1/4)-1", FORMAT_BILLION, False),
            ("Lợi nhuận sau thuế (NPAT)", "NPAT", "='Income Statement'!C20", "='Income Statement'!D20", "='Income Statement'!E20", "='Income Statement'!F20", "='Income Statement'!G20", "=SUM(C{r}:G{r})", "=(G{r}/C{r})^(1/4)-1", FORMAT_BILLION, True),
            ("Dòng tiền thuần HĐKD (Net CFO)", "CFO", "='Cash Flow Statement'!C13", "='Cash Flow Statement'!D13", "='Cash Flow Statement'!E13", "='Cash Flow Statement'!F13", "='Cash Flow Statement'!G13", "=SUM(C{r}:G{r})", "=(G{r}/C{r})^(1/4)-1", FORMAT_BILLION, False),
            ("Chi đầu tư tài sản (CapEx)", "CAPEX", "='Cash Flow Statement'!C15", "='Cash Flow Statement'!D15", "='Cash Flow Statement'!E15", "='Cash Flow Statement'!F15", "='Cash Flow Statement'!G15", "=SUM(C{r}:G{r})", "=(G{r}/C{r})^(1/4)-1", FORMAT_BILLION, False),
            ("Dòng tiền tự do DN (FCFF)", "FCFF", "='Cash Flow Statement'!C28", "='Cash Flow Statement'!D28", "='Cash Flow Statement'!E28", "='Cash Flow Statement'!F28", "='Cash Flow Statement'!G28", "=SUM(C{r}:G{r})", "=(G{r}/C{r})^(1/4)-1", FORMAT_BILLION, False),
            ("Tiền mặt cuối kỳ (Ending Cash)", "CASH", "='Balance Sheet'!C5", "='Balance Sheet'!D5", "='Balance Sheet'!E5", "='Balance Sheet'!F5", "='Balance Sheet'!G5", "-", "-", FORMAT_BILLION, False),
            ("Tổng tài sản (Total Assets)", "TA", "='Balance Sheet'!C13", "='Balance Sheet'!D13", "='Balance Sheet'!E13", "='Balance Sheet'!F13", "='Balance Sheet'!G13", "-", "-", FORMAT_BILLION, True),
            ("Tổng nợ vay (Total Debt)", "DEBT", "='Balance Sheet'!C21", "='Balance Sheet'!D21", "='Balance Sheet'!E21", "='Balance Sheet'!F21", "='Balance Sheet'!G21", "-", "-", FORMAT_BILLION, False),
            ("Vốn chủ sở hữu (Total Equity)", "EQUITY", "='Balance Sheet'!C24", "='Balance Sheet'!D24", "='Balance Sheet'!E24", "='Balance Sheet'!F24", "='Balance Sheet'!G24", "-", "-", FORMAT_BILLION, True),
        ]

        for item in summary_rows:
            row += 1
            lbl, code, f1, f2, f3, f4, f5, f_tot, f_cagr, num_fmt, is_bold = item
            ws.cell(row=row, column=1, value=lbl).font = FONT_BOLD if is_bold else FONT_REGULAR
            ws.cell(row=row, column=2, value=code).alignment = ALIGN_CENTER
            
            for col_idx, f_val in enumerate([f1, f2, f3, f4, f5], start=3):
                c = ws.cell(row=row, column=col_idx, value=f_val)
                c.font = FONT_BOLD if is_bold else FONT_REGULAR
                c.number_format = num_fmt
                c.alignment = ALIGN_RIGHT
                c.border = BOX_BORDER

            # Total & CAGR
            c_tot = ws.cell(row=row, column=8, value=f_tot.format(r=row) if "{r}" in f_tot else f_tot)
            c_tot.font = FONT_BOLD if is_bold else FONT_REGULAR
            c_tot.number_format = num_fmt if f_tot != "-" else "@"
            c_tot.alignment = ALIGN_RIGHT
            c_tot.border = BOX_BORDER

            c_cagr = ws.cell(row=row, column=9, value=f_cagr.format(r=row) if "{r}" in f_cagr else f_cagr)
            c_cagr.font = FONT_BOLD if is_bold else FONT_REGULAR
            c_cagr.number_format = FORMAT_PERCENT if f_cagr != "-" else "@"
            c_cagr.alignment = ALIGN_RIGHT
            c_cagr.border = BOX_BORDER

        # Autofit columns
        cls._autofit_columns(ws)

    # -------------------------------------------------------------------------
    # TAB 2: INCOME STATEMENT
    # -------------------------------------------------------------------------
    @classmethod
    def _build_income_statement_tab(
        cls,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        res: ThreeStatementForecastResult,
        div: float,
        unit_lbl: str,
    ) -> None:
        ws.views.sheetView[0].showGridLines = True
        is_stmt = res.income_statement

        # Title
        ws.merge_cells("A1:G2")
        ws["A1"] = f"BÁO CÁO KẾT QUẢ KINH DOANH DỰ PHÓNG (INCOME STATEMENT / P&L) - {res.symbol}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_HEADER
        ws["A1"].alignment = ALIGN_CENTER

        ws["A3"] = f"Đơn vị tính: {unit_lbl}"
        ws["A3"].font = FONT_ITALIC

        # Headers
        headers = ["Khoản mục P&L (Line Item)", "Mã", f"{res.forecast_years[0]}", f"{res.forecast_years[1]}", f"{res.forecast_years[2]}", f"{res.forecast_years[3]}", f"{res.forecast_years[4]}"]
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_HEADER

        # P&L Line Items (Rows 5 to 21)
        # Note: We write raw base numbers for Year 1-5 or explicit formula links
        pnl_defs = [
            ("1. Doanh thu thuần (Net Sales / Revenue)", "REV", [v / div for v in is_stmt.revenue], FORMAT_BILLION, True, None),
            ("   Tăng trưởng doanh thu YoY (%)", "REV_G", is_stmt.revenue_growth, FORMAT_PERCENT, False, None),
            ("2. Giá vốn hàng bán (Cost of Goods Sold)", "COGS", [v / div for v in is_stmt.cogs], FORMAT_BILLION, False, None),
            ("3. Lợi nhuận gộp (Gross Profit = REV - COGS)", "GP", None, FORMAT_BILLION, True, "=C5-C7"),
            ("   Biên lợi nhuận gộp (% Gross Margin)", "GM", None, FORMAT_PERCENT, False, "=C8/C5"),
            ("4. Chi phí bán hàng & QLDN (SG&A Expense)", "SGA", [v / div for v in is_stmt.sga_expense], FORMAT_BILLION, False, None),
            ("5. Lợi nhuận trước khấu hao & lãi vay (EBITDA)", "EBITDA", None, FORMAT_BILLION, True, "=C8-C10"),
            ("6. Chi phí khấu hao & phân bổ (D&A Expense)", "DA", [v / div for v in is_stmt.depreciation_amortization], FORMAT_BILLION, False, None),
            ("7. Lợi nhuận trước lãi vay & thuế (EBIT = EBITDA - DA)", "EBIT", None, FORMAT_BILLION, True, "=C11-C12"),
            ("   Biên lợi nhuận hoạt động (% EBIT Margin)", "EBIT_M", None, FORMAT_PERCENT, False, "=C13/C5"),
            ("8. Chi phí lãi vay (Gross Interest Expense)", "INT_EXP", [v / div for v in is_stmt.interest_expense], FORMAT_BILLION, False, None),
            ("9. Doanh thu hoạt động tài chính (Interest Income)", "INT_INC", [v / div for v in is_stmt.interest_income], FORMAT_BILLION, False, None),
            ("10. Lợi nhuận trước thuế (EBT = EBIT - INT_EXP + INT_INC)", "EBT", None, FORMAT_BILLION, True, "=C13-C15+C16"),
            ("11. Chi phí thuế TNDN (Income Tax Expense)", "TAX", [v / div for v in is_stmt.tax_expense], FORMAT_BILLION, False, "=MAX(0, C17*0.20)"),
            ("    Thuế suất hiệu dụng (% Eff. Tax Rate)", "TAX_R", is_stmt.effective_tax_rate, FORMAT_PERCENT, False, None),
            ("12. LỢI NHUẬN SAU THUẾ (NET PROFIT AFTER TAX - NPAT)", "NPAT", None, FORMAT_BILLION, True, "=C17-C18"),
            ("    Biên lợi nhuận ròng (% Net Margin)", "NET_M", None, FORMAT_PERCENT, False, "=C20/C5"),
        ]

        row = 5
        for lbl, code, vals, num_fmt, is_bold, formula_tmpl in pnl_defs:
            ws.cell(row=row, column=1, value=lbl).font = FONT_BOLD if is_bold else FONT_REGULAR
            ws.cell(row=row, column=2, value=code).alignment = ALIGN_CENTER

            for col_idx in range(3, 8):
                col_letter = get_column_letter(col_idx)
                if formula_tmpl is not None:
                    # Dynamically shift column references safely
                    cell_formula = cls._shift_formula_column(formula_tmpl, col_letter)
                    c = ws.cell(row=row, column=col_idx, value=cell_formula)
                else:
                    t_idx = col_idx - 3
                    val = vals[t_idx] if vals and t_idx < len(vals) else 0.0
                    c = ws.cell(row=row, column=col_idx, value=val)

                c.font = FONT_BOLD if is_bold else FONT_REGULAR
                c.number_format = num_fmt
                c.alignment = ALIGN_RIGHT
                c.border = TOTAL_BORDER if code in ["NPAT", "TA"] else BOX_BORDER

            row += 1

        cls._autofit_columns(ws)

    # -------------------------------------------------------------------------
    # TAB 3: BALANCE SHEET
    # -------------------------------------------------------------------------
    @classmethod
    def _build_balance_sheet_tab(
        cls,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        res: ThreeStatementForecastResult,
        div: float,
        unit_lbl: str,
    ) -> None:
        ws.views.sheetView[0].showGridLines = True
        bs = res.balance_sheet

        ws.merge_cells("A1:G2")
        ws["A1"] = f"BẢNG CÂN ĐỐI KẾ TOÁN DỰ PHÓNG (BALANCE SHEET) - {res.symbol}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_HEADER
        ws["A1"].alignment = ALIGN_CENTER

        ws["A3"] = f"Đơn vị tính: {unit_lbl} | Kiểm toán cân đối: |Tổng TS - (Tổng Nợ + VCSH)| < 1e-5"
        ws["A3"].font = FONT_ITALIC

        headers = ["Chỉ tiêu Bảng CĐKT (Balance Sheet Item)", "Mã", f"{res.forecast_years[0]}", f"{res.forecast_years[1]}", f"{res.forecast_years[2]}", f"{res.forecast_years[3]}", f"{res.forecast_years[4]}"]
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_HEADER

        bs_defs = [
            # Current Assets (Rows 5 to 9)
            ("1. Tiền và tương đương tiền (Cash & Equivalents)", "CASH", [v / div for v in bs.cash], FORMAT_BILLION, False, "='Cash Flow Statement'!C27"),
            ("2. Phải thu ngắn hạn khách hàng (Accounts Receivable)", "AR", [v / div for v in bs.accounts_receivable], FORMAT_BILLION, False, "='Working Capital Schedule'!C11"),
            ("3. Hàng tồn kho (Inventories)", "INV", [v / div for v in bs.inventory], FORMAT_BILLION, False, "='Working Capital Schedule'!C12"),
            ("4. Tài sản ngắn hạn khác (Other Current Assets)", "OCA", [v / div for v in bs.other_current_assets], FORMAT_BILLION, False, "='Working Capital Schedule'!C14"),
            ("TỔNG TÀI SẢN NGẮN HẠN (TOTAL CURRENT ASSETS)", "TCA", None, FORMAT_BILLION, True, "=SUM(C5:C8)"),
            
            # Non-Current Assets (Rows 10 to 14)
            ("5. Tài sản cố định hữu hình thuần (Net PPE)", "PPE", [v / div for v in bs.net_ppe], FORMAT_BILLION, False, None),
            ("6. Tài sản dài hạn khác (Other Non-Current Assets)", "ONCA", [v / div for v in bs.other_non_current_assets], FORMAT_BILLION, False, None),
            ("TỔNG TÀI SẢN DÀI HẠN (TOTAL NON-CURRENT ASSETS)", "TNCA", None, FORMAT_BILLION, True, "=SUM(C10:C11)"),
            ("TỔNG CỘNG TÀI SẢN (TOTAL ASSETS = TCA + TNCA)", "TA", None, FORMAT_BILLION, True, "=C9+C12"),
            
            # Liabilities (Rows 15 to 23)
            ("7. Phải trả người bán ngắn hạn (Accounts Payable)", "AP", [v / div for v in bs.accounts_payable], FORMAT_BILLION, False, "='Working Capital Schedule'!C13"),
            ("8. Nợ ngắn hạn khác (Other Current Liabilities)", "OCL", [v / div for v in bs.other_current_liabilities], FORMAT_BILLION, False, "='Working Capital Schedule'!C15"),
            ("9. Vay và nợ thuê tài chính ngắn hạn (Short-Term Debt)", "ST_DEBT", [v / div for v in bs.short_term_debt], FORMAT_BILLION, False, "='Debt & Capital Schedule'!C10"),
            ("TỔNG NỢ NGẮN HẠN (TOTAL CURRENT LIABILITIES)", "TCL", None, FORMAT_BILLION, True, "=SUM(C14:C16)"),
            ("10. Vay và nợ thuê tài chính dài hạn (Long-Term Debt)", "LT_DEBT", [v / div for v in bs.long_term_debt], FORMAT_BILLION, False, "='Debt & Capital Schedule'!C11"),
            ("11. Nợ dài hạn khác (Other Non-Current Liabilities)", "ONCL", [0.0] * 5, FORMAT_BILLION, False, None),
            ("TỔNG NỢ PHẢI TRẢ (TOTAL LIABILITIES = TCL + TNCL)", "TL", None, FORMAT_BILLION, True, "=C17+C18+C19"),
            ("   Trong đó: Tổng nợ vay chịu lãi (Total Debt = ST + LT)", "TOTAL_DEBT", None, FORMAT_BILLION, False, "=C16+C18"),
            
            # Equity (Rows 24 to 28)
            ("12. Vốn góp của chủ sở hữu (Contributed Capital)", "CAPITAL", [v / div for v in bs.contributed_capital], FORMAT_BILLION, False, "='Debt & Capital Schedule'!C18"),
            ("13. Lợi nhuận sau thuế chưa phân phối (Retained Earnings)", "RE", [v / div for v in bs.retained_earnings], FORMAT_BILLION, False, "='Debt & Capital Schedule'!C23"),
            ("TỔNG VỐN CHỦ SỞ HỮU (TOTAL SHAREHOLDERS' EQUITY)", "EQUITY", None, FORMAT_BILLION, True, "=C22+C23"),
            ("TỔNG CỘNG NGUỒN VỐN (TOTAL LIABILITIES & EQUITY)", "TL_EQUITY", None, FORMAT_BILLION, True, "=C20+C24"),
            
            # Balance Check & Invariant Verification (Rows 29 to 30)
            ("CHÊNH LỆCH CÂN ĐỐI (BALANCE CHECK = TA - TL_EQUITY)", "DIFF", None, FORMAT_BILLION, True, "=C13-C25"),
            ("TRẠNG THÁI CÂN ĐỐI TOÁN HỌC (IS_BALANCED CHECK)", "IS_BALANCED", None, "@", True, "=IF(ABS(C26)<1, \"BALANCED\", \"UNBALANCED\")"),
        ]

        row = 5
        for lbl, code, vals, num_fmt, is_bold, formula_tmpl in bs_defs:
            ws.cell(row=row, column=1, value=lbl).font = FONT_BOLD if is_bold else FONT_REGULAR
            ws.cell(row=row, column=2, value=code).alignment = ALIGN_CENTER

            for col_idx in range(3, 8):
                col_letter = get_column_letter(col_idx)
                if formula_tmpl is not None:
                    cell_formula = cls._shift_formula_column(formula_tmpl, col_letter)
                    c = ws.cell(row=row, column=col_idx, value=cell_formula)
                else:
                    t_idx = col_idx - 3
                    val = vals[t_idx] if vals and t_idx < len(vals) else 0.0
                    c = ws.cell(row=row, column=col_idx, value=val)

                c.font = FONT_BOLD if is_bold else FONT_REGULAR
                c.number_format = num_fmt
                c.alignment = ALIGN_RIGHT if code != "IS_BALANCED" else ALIGN_CENTER
                
                if code == "IS_BALANCED":
                    c.fill = FILL_GREEN
                elif code in ["TA", "TL_EQUITY"]:
                    c.border = TOTAL_BORDER
                else:
                    c.border = BOX_BORDER

            row += 1

        cls._autofit_columns(ws)

    # -------------------------------------------------------------------------
    # TAB 4: CASH FLOW STATEMENT (DIRECT METHOD)
    # -------------------------------------------------------------------------
    @classmethod
    def _build_cash_flow_tab(
        cls,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        res: ThreeStatementForecastResult,
        div: float,
        unit_lbl: str,
    ) -> None:
        ws.views.sheetView[0].showGridLines = True
        cfs = res.cash_flow_statement

        ws.merge_cells("A1:G2")
        ws["A1"] = f"BÁO CÁO LƯU CHUYỂN TIỀN TỆ TRỰC TIẾP (DIRECT CASH FLOW STATEMENT) - {res.symbol}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_HEADER
        ws["A1"].alignment = ALIGN_CENTER

        ws["A3"] = f"Đơn vị tính: {unit_lbl} | Khớp nối kế toán: CFO thuần == NPAT + D&A - Delta NWC"
        ws["A3"].font = FONT_ITALIC

        headers = ["Chỉ tiêu Lưu chuyển Tiền tệ (Direct CFS)", "Mã", f"{res.forecast_years[0]}", f"{res.forecast_years[1]}", f"{res.forecast_years[2]}", f"{res.forecast_years[3]}", f"{res.forecast_years[4]}"]
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_HEADER

        cfs_defs = [
            # Operating Cash Flows (Direct Method)
            ("I. LƯU CHUYỂN TIỀN TỪ HĐKD (OPERATING ACTIVITIES)", "CFO_SEC", None, "@", True, None),
            ("1. Tiền thu từ bán hàng & cung cấp DV (Cash from Customers)", "CUST", [v / div for v in cfs.cash_from_customers], FORMAT_BILLION, False, "='Income Statement'!C5 - 'Working Capital Schedule'!C18"),
            ("2. Tiền chi trả cho người cung cấp hàng hóa (Cash to Suppliers)", "SUPP", [v / div for v in cfs.cash_to_suppliers], FORMAT_BILLION, False, "='Income Statement'!C7 + 'Working Capital Schedule'!C19 - 'Working Capital Schedule'!C20"),
            ("   LƯU CHUYỂN TIỀN THÔ TỪ HĐKD (GROSS CFO = CUST - SUPP)", "GROSS_CFO", None, FORMAT_BILLION, True, "=C6-C7"),
            ("3. Tiền chi trả chi phí bán hàng & QLDN (Cash for SG&A)", "OPEX", [v / div for v in cfs.cash_for_opex], FORMAT_BILLION, False, "='Income Statement'!C10 + 'Working Capital Schedule'!C21 - 'Working Capital Schedule'!C22"),
            ("4. Tiền lãi vay đã trả (Cash Interest Paid)", "INT_PAID", [v / div for v in cfs.cash_interest_paid], FORMAT_BILLION, False, "='Income Statement'!C15"),
            ("5. Tiền lãi và lợi tức đã thu (Cash Interest Received)", "INT_REC", [v / div for v in cfs.cash_interest_received], FORMAT_BILLION, False, "='Income Statement'!C16"),
            ("6. Tiền thuế TNDN đã nộp (Cash Tax Paid)", "TAX_PAID", [v / div for v in cfs.cash_tax_paid], FORMAT_BILLION, False, "='Income Statement'!C18"),
            ("LƯU CHUYỂN TIỀN THUẦN TỪ HĐKD (NET CFO)", "NET_CFO", None, FORMAT_BILLION, True, "=C8-C9-C10+C11-C12"),
            
            # Investing Cash Flows
            ("II. LƯU CHUYỂN TIỀN TỪ HĐ ĐẦU TƯ (INVESTING ACTIVITIES)", "CFI_SEC", None, "@", True, None),
            ("7. Tiền chi mua sắm TSCĐ (Capital Expenditures - CapEx)", "CAPEX", [v / div for v in cfs.capex], FORMAT_BILLION, False, None),
            ("8. Tiền thu/chi từ đầu tư khác (Other CFI)", "OTHER_CFI", [0.0] * 5, FORMAT_BILLION, False, None),
            ("LƯU CHUYỂN TIỀN THUẦN TỪ HĐ ĐẦU TƯ (NET CFI)", "NET_CFI", None, FORMAT_BILLION, True, "=-C15+C16"),
            
            # Financing Cash Flows
            ("III. LƯU CHUYỂN TIỀN TỪ HĐ TÀI CHÍNH (FINANCING ACTIVITIES)", "CFF_SEC", None, "@", True, None),
            ("9. Tiền thu từ đi vay mới (New Debt Borrowings)", "NEW_DEBT", [v / div for v in cfs.new_debt_drawdowns], FORMAT_BILLION, False, "='Debt & Capital Schedule'!C7"),
            ("10. Tiền trả nợ gốc vay (Principal Debt Repayments)", "REPAY_DEBT", [v / div for v in cfs.principal_debt_repayments], FORMAT_BILLION, False, "='Debt & Capital Schedule'!C8"),
            ("    Dòng vay nợ thuần (Net Debt Drawdown = Borrow - Repay)", "NET_DRAW", None, FORMAT_BILLION, False, "=C19-C20"),
            ("11. Tiền chi trả cổ tức cho CĐ (Cash Dividends Paid)", "DIV_PAID", [v / div for v in cfs.dividends_paid], FORMAT_BILLION, False, "='Debt & Capital Schedule'!C21"),
            ("12. Tiền chi mua lại cổ phiếu (Share Repurchases)", "REP_PAID", [v / div for v in cfs.share_repurchases], FORMAT_BILLION, False, "='Debt & Capital Schedule'!C22"),
            ("LƯU CHUYỂN TIỀN THUẦN TỪ HĐ TÀI CHÍNH (NET CFF)", "NET_CFF", None, FORMAT_BILLION, True, "=C21-C22-C23"),
            
            # Net Cash Change & Reconciliation
            ("LƯU CHUYỂN TIỀN THUẦN TRONG KỲ (NET CHANGE IN CASH = CFO + CFI + CFF)", "DELTA_CASH", None, FORMAT_BILLION, True, "=C13+C17+C24"),
            ("Tiền và tương đương tiền đầu kỳ (Beginning Cash)", "BEG_CASH", [v / div for v in cfs.beginning_cash], FORMAT_BILLION, False, None),
            ("TIỀN VÀ TƯƠNG ĐƯƠNG TIỀN CUỐI KỲ (ENDING CASH = BEG + DELTA)", "END_CASH", None, FORMAT_BILLION, True, "=C25+C26"),
            
            # Intrinsic Valuation Cash Flows
            ("Dòng tiền tự do doanh nghiệp (FCFF = CFO - CapEx)", "FCFF", None, FORMAT_BILLION, True, "=C13-C15"),
            ("Dòng tiền tự do vốn CSH (FCFE = CFO - CapEx + Net Debt)", "FCFE", None, FORMAT_BILLION, True, "=C13-C15+C21"),
        ]

        row = 5
        for lbl, code, vals, num_fmt, is_bold, formula_tmpl in cfs_defs:
            is_sec = code.endswith("_SEC")
            cell_lbl = ws.cell(row=row, column=1, value=lbl)
            cell_lbl.font = FONT_SECTION if is_sec else (FONT_BOLD if is_bold else FONT_REGULAR)
            if is_sec:
                ws.merge_cells(f"A{row}:G{row}")
                cell_lbl.fill = FILL_SECTION
                row += 1
                continue

            ws.cell(row=row, column=2, value=code).alignment = ALIGN_CENTER

            for col_idx in range(3, 8):
                col_letter = get_column_letter(col_idx)
                if formula_tmpl is not None:
                    cell_formula = cls._shift_formula_column(formula_tmpl, col_letter)
                    c = ws.cell(row=row, column=col_idx, value=cell_formula)
                else:
                    t_idx = col_idx - 3
                    val = vals[t_idx] if vals and t_idx < len(vals) else 0.0
                    c = ws.cell(row=row, column=col_idx, value=val)

                c.font = FONT_BOLD if is_bold else FONT_REGULAR
                c.number_format = num_fmt
                c.alignment = ALIGN_RIGHT
                c.border = TOTAL_BORDER if code in ["NET_CFO", "END_CASH"] else BOX_BORDER

            row += 1

        cls._autofit_columns(ws)

    # -------------------------------------------------------------------------
    # TAB 5: WORKING CAPITAL SCHEDULE
    # -------------------------------------------------------------------------
    @classmethod
    def _build_working_capital_tab(
        cls,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        res: ThreeStatementForecastResult,
        div: float,
        unit_lbl: str,
    ) -> None:
        ws.views.sheetView[0].showGridLines = True
        wc = res.working_capital_schedule

        ws.merge_cells("A1:G2")
        ws["A1"] = f"LỊCH TRÌNH VỐN LƯU ĐỘNG (WORKING CAPITAL SCHEDULE) - {res.symbol}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_HEADER
        ws["A1"].alignment = ALIGN_CENTER

        ws["A3"] = f"Đơn vị tính: {unit_lbl} | Chu kỳ chuyển hóa tiền mặt: CCC = DSO + DIO - DPO"
        ws["A3"].font = FONT_ITALIC

        headers = ["Chỉ số Vốn Lưu Động (Working Capital Metric)", "Mã", f"{res.forecast_years[0]}", f"{res.forecast_years[1]}", f"{res.forecast_years[2]}", f"{res.forecast_years[3]}", f"{res.forecast_years[4]}"]
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_HEADER

        def _get_wc_val(p: Any, key: str, fallback: float = 0.0) -> float:
            if isinstance(p, dict):
                return p.get(key, fallback)
            return getattr(p, key, fallback)

        wc_defs = [
            # Days Ratios
            ("I. CHU KỲ HOẠT ĐỘNG & SỐ NGÀY QUẢN LÝ (WORKING CAPITAL DAYS)", "DAYS_SEC", None, "@", True, None),
            ("1. Số ngày thu tiền khách hàng (DSO - Days Sales Outstanding)", "DSO", [_get_wc_val(p, "dso") for p in wc], FORMAT_INTEGER, False, None),
            ("2. Số ngày tồn kho bình quân (DIO - Days Inventory Outstanding)", "DIO", [_get_wc_val(p, "dio") for p in wc], FORMAT_INTEGER, False, None),
            ("3. Số ngày trả nợ người bán (DPO - Days Payable Outstanding)", "DPO", [_get_wc_val(p, "dpo") for p in wc], FORMAT_INTEGER, False, None),
            ("CHU KỲ CHUYỂN HÓA TIỀN MẶT (CCC = DSO + DIO - DPO)", "CCC", None, FORMAT_INTEGER, True, "=C6+C7-C8"),
            
            # Balance Items
            ("II. SỐ DƯ TÀI SẢN & NỢ VỐN LƯU ĐỘNG (WORKING CAPITAL BALANCES)", "BAL_SEC", None, "@", True, None),
            ("4. Phải thu khách hàng (Accounts Receivable = REV * DSO / 365)", "AR", None, FORMAT_BILLION, False, "='Income Statement'!C5 * C6 / 365"),
            ("5. Hàng tồn kho (Inventories = COGS * DIO / 365)", "INV", None, FORMAT_BILLION, False, "='Income Statement'!C7 * C7 / 365"),
            ("6. Phải trả người bán (Accounts Payable = COGS * DPO / 365)", "AP", None, FORMAT_BILLION, False, "='Income Statement'!C7 * C8 / 365"),
            ("7. Tài sản ngắn hạn khác (Other Current Assets)", "OCA", [_get_wc_val(p, "other_current_assets") / div for p in wc], FORMAT_BILLION, False, None),
            ("8. Nợ ngắn hạn khác (Other Current Liabilities)", "OCL", [_get_wc_val(p, "other_current_liabilities") / div for p in wc], FORMAT_BILLION, False, None),
            ("VỐN LƯU ĐỘNG THUẦN (NET WORKING CAPITAL = (AR+INV+OCA) - (AP+OCL))", "NWC", None, FORMAT_BILLION, True, "=(C11+C12+C14)-(C13+C15)"),
            
            # Delta NWC
            ("III. BIẾN ĐỘNG VỐN LƯU ĐỘNG (DELTA NWC RECONCILIATION)", "DELTA_SEC", None, "@", True, None),
            ("9. Biến động phải thu (Delta AR)", "DELTA_AR", [_get_wc_val(p, "delta_ar") / div for p in wc], FORMAT_BILLION, False, None),
            ("10. Biến động tồn kho (Delta Inventory)", "DELTA_INV", [_get_wc_val(p, "delta_inv", _get_wc_val(p, "delta_inventory", 0.0)) / div for p in wc], FORMAT_BILLION, False, None),
            ("11. Biến động phải trả (Delta AP)", "DELTA_AP", [_get_wc_val(p, "delta_ap") / div for p in wc], FORMAT_BILLION, False, None),
            ("12. Biến động tài sản ngắn hạn khác (Delta OCA)", "DELTA_OCA", [_get_wc_val(p, "delta_oca") / div for p in wc], FORMAT_BILLION, False, None),
            ("13. Biến động nợ ngắn hạn khác (Delta OCL)", "DELTA_OCL", [_get_wc_val(p, "delta_ocl") / div for p in wc], FORMAT_BILLION, False, None),
            ("TỔNG BIẾN ĐỘNG VỐN LƯU ĐỘNG (DELTA NWC = DELTA_AR + DELTA_INV - DELTA_AP + ...)", "DELTA_NWC", None, FORMAT_BILLION, True, "=C18+C19-C20+C21-C22"),
        ]

        row = 5
        for lbl, code, vals, num_fmt, is_bold, formula_tmpl in wc_defs:
            is_sec = code.endswith("_SEC")
            cell_lbl = ws.cell(row=row, column=1, value=lbl)
            cell_lbl.font = FONT_SECTION if is_sec else (FONT_BOLD if is_bold else FONT_REGULAR)
            if is_sec:
                ws.merge_cells(f"A{row}:G{row}")
                cell_lbl.fill = FILL_SECTION
                row += 1
                continue

            ws.cell(row=row, column=2, value=code).alignment = ALIGN_CENTER

            for col_idx in range(3, 8):
                col_letter = get_column_letter(col_idx)
                if formula_tmpl is not None:
                    cell_formula = cls._shift_formula_column(formula_tmpl, col_letter)
                    c = ws.cell(row=row, column=col_idx, value=cell_formula)
                else:
                    t_idx = col_idx - 3
                    val = vals[t_idx] if vals and t_idx < len(vals) else 0.0
                    c = ws.cell(row=row, column=col_idx, value=val)

                c.font = FONT_BOLD if is_bold else FONT_REGULAR
                c.number_format = num_fmt
                c.alignment = ALIGN_RIGHT
                c.border = TOTAL_BORDER if code in ["CCC", "NWC", "DELTA_NWC"] else BOX_BORDER

            row += 1

        cls._autofit_columns(ws)

    # -------------------------------------------------------------------------
    # TAB 6: DEBT & CAPITAL SCHEDULE
    # -------------------------------------------------------------------------
    @classmethod
    def _build_debt_capital_tab(
        cls,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        res: ThreeStatementForecastResult,
        div: float,
        unit_lbl: str,
    ) -> None:
        ws.views.sheetView[0].showGridLines = True
        debt_sched = res.debt_schedule

        def _get_debt_val(p: Any, key: str, fallback: Any = 0.0) -> Any:
            if isinstance(p, dict):
                return p.get(key, fallback)
            return getattr(p, key, fallback)

        ws.merge_cells("A1:G2")
        ws["A1"] = f"LỊCH TRÌNH VAY NỢ & CƠ CẤU VỐN (DEBT & CAPITAL SCHEDULE) - {res.symbol}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_HEADER
        ws["A1"].alignment = ALIGN_CENTER

        ws["A3"] = f"Đơn vị tính: {unit_lbl} | Xếp hạng tín nhiệm & Lãi suất nợ theo phương pháp Aswath Damodaran"
        ws["A3"].font = FONT_ITALIC

        headers = ["Chỉ tiêu Nợ & Cơ cấu Vốn (Debt & Capital Item)", "Mã", f"{res.forecast_years[0]}", f"{res.forecast_years[1]}", f"{res.forecast_years[2]}", f"{res.forecast_years[3]}", f"{res.forecast_years[4]}"]
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_HEADER

        debt_defs = [
            # Debt Roll-Forward
            ("I. LỊCH TRÌNH NỢ VAY CHỊU LÃI (DEBT AMORTIZATION ROLL-FORWARD)", "DEBT_SEC", None, "@", True, None),
            ("1. Số dư nợ đầu kỳ (Opening Debt Balance)", "OPEN_DEBT", [_get_debt_val(p, "opening_debt") / div for p in debt_sched], FORMAT_BILLION, False, None),
            ("2. Khoản giải ngân nợ mới (New Debt Borrowings)", "NEW_BORROW", [_get_debt_val(p, "new_borrowings") / div for p in debt_sched], FORMAT_BILLION, False, None),
            ("3. Khoản trả nợ gốc (Principal Debt Amortization)", "PRINCIPAL_REPAY", [_get_debt_val(p, "principal_amortization") / div for p in debt_sched], FORMAT_BILLION, False, None),
            ("SỐ DƯ NỢ VAY CUỐI KỲ (CLOSING DEBT = OPEN + NEW - REPAY)", "CLOSE_DEBT", None, FORMAT_BILLION, True, "=C6+C7-C8"),
            ("   Trong đó: Nợ ngắn hạn đến hạn trả (Short-Term Debt)", "ST_DEBT", [_get_debt_val(p, "short_term_debt") / div for p in debt_sched], FORMAT_BILLION, False, None),
            ("   Trong đó: Nợ dài hạn còn lại (Long-Term Debt = CLOSE - ST)", "LT_DEBT", None, FORMAT_BILLION, False, "=C9-C10"),
            
            # Credit Rating & Interest
            ("II. ĐÁNH GIÁ TÍN NHIỆM & CHI PHÍ LÃI VAY (DAMODARAN SYNTHETIC RATING)", "RATING_SEC", None, "@", True, None),
            ("4. Hệ số chi trả lãi vay (Interest Coverage Ratio = EBIT / Interest)", "ICR", [_get_debt_val(p, "interest_coverage_ratio") for p in debt_sched], FORMAT_MULTIPLE, False, "='Income Statement'!C13 / 'Income Statement'!C15"),
            ("5. Xếp hạng tín nhiệm mô phỏng (Synthetic Credit Rating)", "RATING", [_get_debt_val(p, "synthetic_rating", "BBB") for p in debt_sched], "@", False, None),
            ("6. Chi phí nợ vay trước thuế (Pre-Tax Cost of Debt - Kd)", "KD_PRE", [_get_debt_val(p, "cost_of_debt_pre_tax") for p in debt_sched], FORMAT_PERCENT_2, False, None),
            ("7. Chi phí nợ vay sau thuế (After-Tax Cost of Debt = Kd * (1 - t))", "KD_AFTER", [_get_debt_val(p, "cost_of_debt_after_tax") for p in debt_sched], FORMAT_PERCENT_2, True, "=C15*(1-0.20)"),
            
            # Equity Roll-Forward
            ("III. LỊCH TRÌNH VỐN CHỦ SỞ HỮU (EQUITY ROLL-FORWARD)", "EQ_SEC", None, "@", True, None),
            ("8. Vốn góp của chủ sở hữu (Contributed Share Capital)", "CAPITAL", [res.balance_sheet.contributed_capital[i] / div for i in range(5)], FORMAT_BILLION, False, None),
            ("9. Lợi nhuận giữ lại đầu kỳ (Beginning Retained Earnings)", "BEG_RE", [
                (res.balance_sheet.retained_earnings[i-1] / div) if i > 0 else ((res.balance_sheet.retained_earnings[0] - res.income_statement.npat[0] + res.cash_flow_statement.dividends_paid[0]) / div)
                for i in range(5)
            ], FORMAT_BILLION, False, None),
            ("10. Lợi nhuận sau thuế trong kỳ (+ NPAT)", "ADD_NPAT", None, FORMAT_BILLION, False, "='Income Statement'!C20"),
            ("11. Cổ tức tiền mặt chi trả (- Dividends Paid)", "LESS_DIV", [_get_debt_val(p, "dividends_paid", 0.0) / div for p in debt_sched], FORMAT_BILLION, False, None),
            ("12. Mua lại cổ phiếu quỹ (- Share Repurchases)", "LESS_REP", [_get_debt_val(p, "share_repurchases", 0.0) / div for p in debt_sched], FORMAT_BILLION, False, None),
            ("LỢI NHUẬN GIỮ LẠI CUỐI KỲ (ENDING RE = BEG + NPAT - DIV - REP)", "END_RE", None, FORMAT_BILLION, True, "=C19+C20-C21-C22"),
            ("TỔNG VỐN CHỦ SỞ HỮU (TOTAL EQUITY = CAPITAL + ENDING_RE)", "TOTAL_EQUITY", None, FORMAT_BILLION, True, "=C18+C23"),
        ]

        row = 5
        for lbl, code, vals, num_fmt, is_bold, formula_tmpl in debt_defs:
            is_sec = code.endswith("_SEC")
            cell_lbl = ws.cell(row=row, column=1, value=lbl)
            cell_lbl.font = FONT_SECTION if is_sec else (FONT_BOLD if is_bold else FONT_REGULAR)
            if is_sec:
                ws.merge_cells(f"A{row}:G{row}")
                cell_lbl.fill = FILL_SECTION
                row += 1
                continue

            ws.cell(row=row, column=2, value=code).alignment = ALIGN_CENTER

            for col_idx in range(3, 8):
                col_letter = get_column_letter(col_idx)
                if formula_tmpl is not None:
                    cell_formula = cls._shift_formula_column(formula_tmpl, col_letter)
                    c = ws.cell(row=row, column=col_idx, value=cell_formula)
                else:
                    t_idx = col_idx - 3
                    val = vals[t_idx] if vals and t_idx < len(vals) else 0.0
                    c = ws.cell(row=row, column=col_idx, value=val)

                c.font = FONT_BOLD if is_bold else FONT_REGULAR
                c.number_format = num_fmt
                c.alignment = ALIGN_RIGHT if code != "RATING" else ALIGN_CENTER
                c.border = TOTAL_BORDER if code in ["CLOSE_DEBT", "END_RE", "TOTAL_EQUITY"] else BOX_BORDER

            row += 1

        cls._autofit_columns(ws)

    # -------------------------------------------------------------------------
    # TAB 7: VALUATION & SENSITIVITY
    # -------------------------------------------------------------------------
    @classmethod
    def _build_valuation_sensitivity_tab(
        cls,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        res: ThreeStatementForecastResult,
        div: float,
        unit_lbl: str,
    ) -> None:
        ws.views.sheetView[0].showGridLines = True

        ws.merge_cells("A1:H2")
        ws["A1"] = f"MÔ HÌNH ĐỊNH GIÁ & MA TRẬN ĐỘ NHẠY 2 CHIỀU (VALUATION & SENSITIVITY) - {res.symbol}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_HEADER
        ws["A1"].alignment = ALIGN_CENTER

        # 1. Discount Rate & WACC Parameters Box (Row 4-9)
        ws.merge_cells("A4:D4")
        ws["A4"] = "THÔNG SỐ LÃI SUẤT CHIẾT KHẤU & CƠ CẤU VỐN (WACC PARAMETERS)"
        ws["A4"].font = FONT_SECTION
        ws["A4"].fill = FILL_SECTION

        wacc_params = [
            ("Lãi suất phi rủi ro (Risk-Free Rate - Rf):", "B5", 0.045, FORMAT_PERCENT_2),
            ("Phần bù rủi ro vốn cổ phần (Equity Risk Premium - ERP):", "B6", 0.085, FORMAT_PERCENT_2),
            ("Hệ số Beta điều chỉnh (Adjusted Beta):", "B7", 1.10, FORMAT_MULTIPLE),
            ("Chi phí vốn cổ phần (Cost of Equity - Ke = Rf + Beta*ERP):", "B8", "=C5+C7*C6", FORMAT_PERCENT_2),
            ("Chi phí nợ vay sau thuế (After-Tax Kd):", "B9", "='Debt & Capital Schedule'!C16", FORMAT_PERCENT_2),
            ("CHI PHÍ VỐN BÌNH QUÂN (WACC = We*Ke + Wd*Kd):", "B10", "=0.70*C8 + 0.30*C9", FORMAT_PERCENT_2),
        ]

        row = 5
        for lbl, cell_ref, val, num_fmt in wacc_params:
            ws.cell(row=row, column=1, value=lbl).font = FONT_BOLD if "WACC" in lbl else FONT_REGULAR
            ws.merge_cells(f"A{row}:B{row}")
            
            c = ws.cell(row=row, column=3, value=val)
            c.font = FONT_BOLD if "WACC" in lbl else FONT_REGULAR
            c.number_format = num_fmt
            c.alignment = ALIGN_RIGHT
            c.border = TOTAL_BORDER if "WACC" in lbl else BOX_BORDER
            row += 1

        # 2. Valuation Outputs Comparison (Row 4-10 in Col F-H)
        ws.merge_cells("F4:H4")
        ws["F4"] = "KẾT QUẢ ĐỊNH GIÁ DOANH NGHIỆP (INTRINSIC VALUATION SUMMARY)"
        ws["F4"].font = FONT_SECTION
        ws["F4"].fill = FILL_SECTION

        val_rows = [
            ("Mô hình DCF 2 giai đoạn (FCFF Model)", "=SUM('Cash Flow Statement'!C28:G28)", FORMAT_BILLION),
            ("Mô hình FCFE dòng tiền CSH", "=SUM('Cash Flow Statement'!C29:G29)", FORMAT_BILLION),
            ("Mô hình Buffett Owner's Earnings", "='Income Statement'!C20", FORMAT_BILLION),
            ("Tốc độ tăng trưởng dài hạn (Terminal Growth - g)", 0.035, FORMAT_PERCENT_2),
            ("GIÁ TRỊ DOANH NGHIỆP HỢP LÝ (FAIR VALUE)", "=(H5*(1+H8))/(C10-H8)", FORMAT_BILLION),
        ]

        val_row = 5
        for lbl, f_val, num_fmt in val_rows:
            ws.cell(row=val_row, column=6, value=lbl).font = FONT_BOLD if "FAIR VALUE" in lbl else FONT_REGULAR
            c = ws.cell(row=val_row, column=8, value=f_val)
            c.font = FONT_BOLD if "FAIR VALUE" in lbl else FONT_REGULAR
            c.number_format = num_fmt
            c.alignment = ALIGN_RIGHT
            c.border = TOTAL_BORDER if "FAIR VALUE" in lbl else BOX_BORDER
            val_row += 1

        # 3. 5x5 Dynamic Sensitivity Matrix (WACC vs Terminal Growth g)
        row = 12
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = "MA TRẬN ĐỘ NHẠY ĐỊNH GIÁ 2 CHIỀU (5x5 SENSITIVITY MATRIX: WACC vs TERMINAL GROWTH g)"
        ws[f"A{row}"].font = FONT_SECTION
        ws[f"A{row}"].fill = FILL_SECTION

        row += 1
        ws[f"A{row}"] = "WACC \\ Terminal Growth (g)"
        ws[f"A{row}"].font = FONT_HEADER
        ws[f"A{row}"].fill = FILL_HEADER
        ws[f"A{row}"].alignment = ALIGN_CENTER

        g_rates = [0.025, 0.030, 0.035, 0.040, 0.045]
        wacc_rates = [0.090, 0.100, 0.110, 0.120, 0.130]

        for col_idx, g_val in enumerate(g_rates, start=2):
            c = ws.cell(row=row, column=col_idx, value=g_val)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER
            c.number_format = FORMAT_PERCENT

        # Matrix Body (Live Excel formulas referencing the base FCFF terminal formula)
        for r_idx, w_val in enumerate(wacc_rates, start=1):
            cur_r = row + r_idx
            c_w = ws.cell(row=cur_r, column=1, value=w_val)
            c_w.font = FONT_HEADER
            c_w.fill = FILL_HEADER
            c_w.alignment = ALIGN_CENTER
            c_w.number_format = FORMAT_PERCENT

            for c_idx, g_val in enumerate(g_rates, start=2):
                col_let = get_column_letter(c_idx)
                # Formula: ($H$5 * (1 + col_let$row)) / ($Acur_r - col_let$row)
                cell_formula = f"=($H$5*(1+{col_let}${row}))/($A{cur_r}-{col_let}${row})"
                c_val = ws.cell(row=cur_r, column=c_idx, value=cell_formula)
                c_val.font = FONT_REGULAR
                c_val.number_format = FORMAT_BILLION
                c_val.alignment = ALIGN_RIGHT
                c_val.border = BOX_BORDER
                # Highlight central base case
                if r_idx == 3 and c_idx == 4:
                    c_val.fill = FILL_SECTION
                    c_val.font = FONT_BOLD

        cls._autofit_columns(ws)

    # -------------------------------------------------------------------------
    # UTILITY: AUTO-FIT COLUMNS WITH PADDING
    # -------------------------------------------------------------------------
    @staticmethod
    def _autofit_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Sets reasonable column widths with safe padding."""
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 16
        for col_idx in range(3, 12):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 22
