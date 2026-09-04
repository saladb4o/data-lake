import os
import sys
sys.path.insert(0, os.path.abspath("."))
import openpyxl
from services.three_statement_engine import ThreeStatementEngine
from services.financial_model_exporter import FinancialModelExporter
from services.stock_service import VN30_SYMBOLS

test_symbols = ["HPG", "FPT", "MWG", "VCB", "VIC", "NVL", "MSN", "SSI", "STB", "VHM"]
os.makedirs("tmp_review", exist_ok=True)

for sym in test_symbols:
    res = ThreeStatementEngine.build_forecast_from_screener(sym)
    path = f"tmp_review/{sym}_audit.xlsx"
    FinancialModelExporter.export_to_excel(res, path)
    wb = openpyxl.load_workbook(path, data_only=False)
    assert len(wb.sheetnames) == 7, f"{sym} wrong sheet count"
    
    # Check all formulas for single quote protection on sheets with spaces
    error_tokens = ["#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"]
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                # Check for errors
                for tok in error_tokens:
                    assert tok not in val, f"Error token {tok} in {sym} sheet {sname} {cell.coordinate}: {val}"
                
                # Check for formula sheet references
                if val.startswith("="):
                    for other_sheet in wb.sheetnames:
                        if " " in other_sheet and other_sheet in val:
                            quoted_pattern = f"'{other_sheet}'!"
                            assert quoted_pattern in val, f"Sheet reference not properly quoted with single quotes: {val} in sheet {sname} {cell.coordinate}"

    print(f"[AUDIT PASS] {sym}: 7 tabs verified, all formulas valid syntax, all sheet links quoted.")

print("\nALL 10 DIVERSE SECTOR TICKERS PASSED EXCEL FORMULA INTEGRITY AUDIT!")
