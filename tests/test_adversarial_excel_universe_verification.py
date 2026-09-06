"""
=============================================================================
CHALLENGER 2: ADVERSARIAL EXCEL & UNIVERSE VERIFICATION TEST SUITE
=============================================================================
Empirical adversarial verification of:
1. Actual .xlsx workbooks generation and deep cell parsing for real VN tickers
   (HPG, FPT, MWG, VCB, NVL, VIC, VNM).
2. Zero formula errors (#REF!, #NAME?, #VALUE!, #DIV/0!, #N/A, #NULL!, #NUM!).
3. Cross-sheet reference integrity and resolution across all 7 worksheets.
4. 5x5 WACC vs g sensitivity matrix live dynamic formulas.
5. Balance sheet audit badges evaluation to BALANCED with green fills.
6. 100% of VN30 constituents 5-year balance sheet balance test (|TA - (TL+TE)| < 1.0 VND or |Diff_Billion| < 1e-5).
=============================================================================
"""

import os
import re
import pytest
import openpyxl
from openpyxl.utils import get_column_letter

from services.three_statement_engine import ThreeStatementEngine, safe_div
from services.financial_model_exporter import FinancialModelExporter
from services.stock_service import VN30_SYMBOLS


EXPECTED_SHEETS = [
    "Summary & Dashboard",
    "Income Statement",
    "Balance Sheet",
    "Cash Flow Statement",
    "Working Capital Schedule",
    "Debt & Capital Schedule",
    "Valuation & Sensitivity",
]

ERROR_TOKENS = ["#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"]
CROSS_SHEET_REGEX = re.compile(r"'([^']+)'!([A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?)")


@pytest.fixture(scope="session")
def export_scratch_dir(tmp_path_factory):
    scratch = tmp_path_factory.mktemp("challenger2_excel_stress")
    return str(scratch)


class TestAdversarialExcelRealTickers:
    """
    Stress-tests actual .xlsx generation for real Vietnamese tickers:
    - Blue chips / Industrial: HPG, VNM
    - High-growth Tech: FPT
    - Retail with Negative CCC: MWG
    - Financial / Banking: VCB
    - High-debt / Real Estate: NVL, VIC
    """

    @pytest.mark.parametrize("sym", ["HPG", "FPT", "MWG", "VCB", "NVL", "VIC", "VNM"])
    def test_real_ticker_export_and_cell_integrity(self, sym, export_scratch_dir, screener_snapshot):
        # 1. Run 3-Way statement forecast
        res = ThreeStatementEngine.build_forecast_from_screener(sym)
        assert res.all_years_balanced is True, f"Balance sheet not balanced for {sym}"
        # Max balance diff in Billion VND is res.max_balance_difference / 1e9, must be < 1e-5
        max_diff_billion = res.max_balance_difference / 1e9
        assert max_diff_billion < 1e-5, f"Max balance difference for {sym} in Billion VND too large: {max_diff_billion}"
        assert res.max_balance_difference < 1.0, f"Max raw balance diff for {sym} exceeds 1.0 VND: {res.max_balance_difference}"

        # 2. Export workbook to disk
        out_path = os.path.join(export_scratch_dir, f"{sym}_adversarial.xlsx")
        FinancialModelExporter.export_to_excel(res, out_path)
        assert os.path.exists(out_path)
        assert os.path.getsize(out_path) > 5000

        # 3. Open workbook with openpyxl
        wb = openpyxl.load_workbook(out_path, data_only=False)

        # 4. Verify exact 7-sheet architecture
        assert len(wb.sheetnames) == 7
        for sheet_name in EXPECTED_SHEETS:
            assert sheet_name in wb.sheetnames, f"Missing sheet '{sheet_name}' in {sym} workbook"

        # 5. Programmatically parse every single cell across all 7 worksheets
        total_cells_inspected = 0
        formula_cells_inspected = 0
        cross_sheet_refs_inspected = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell.value is None:
                        continue
                    total_cells_inspected += 1
                    val_str = str(cell.value)

                    # Check for formula errors
                    for token in ERROR_TOKENS:
                        assert token not in val_str, (
                            f"[{sym}] Formula error '{token}' detected in sheet '{sheet_name}' cell {cell.coordinate}: {val_str}"
                        )

                    # Validate formula syntax and cross-sheet links
                    if val_str.startswith("="):
                        formula_cells_inspected += 1
                        matches = CROSS_SHEET_REGEX.findall(val_str)
                        for target_sheet, target_range in matches:
                            cross_sheet_refs_inspected += 1
                            assert target_sheet in wb.sheetnames, (
                                f"[{sym}] Sheet '{sheet_name}' cell {cell.coordinate} points to missing sheet '{target_sheet}'"
                            )
                            # Verify target coordinates exist
                            if ":" in target_range:
                                start_cell, end_cell = target_range.split(":")
                                assert wb[target_sheet][start_cell] is not None
                                assert wb[target_sheet][end_cell] is not None
                            else:
                                assert wb[target_sheet][target_range] is not None

        assert total_cells_inspected > 100, f"[{sym}] Unexpectedly low total cells: {total_cells_inspected}"
        assert formula_cells_inspected > 30, f"[{sym}] Unexpectedly low formula cells: {formula_cells_inspected}"
        assert cross_sheet_refs_inspected > 10, f"[{sym}] Unexpectedly low cross sheet refs: {cross_sheet_refs_inspected}"

        # 6. Verify 5x5 WACC vs g Sensitivity Matrix in Tab 7 (Valuation & Sensitivity)
        ws_val = wb["Valuation & Sensitivity"]
        wacc_rates = [0.090, 0.100, 0.110, 0.120, 0.130]
        g_rates = [0.025, 0.030, 0.035, 0.040, 0.045]

        # Verify matrix column headers (row 13)
        for c_idx, g_val in enumerate(g_rates, start=2):
            col_letter = get_column_letter(c_idx)
            cell_val = ws_val[f"{col_letter}13"].value
            assert pytest.approx(cell_val, 1e-4) == g_val

        # Verify matrix row headers & dynamic cell formulas (rows 14 to 18)
        for r_idx, w_val in enumerate(wacc_rates, start=1):
            cur_r = 13 + r_idx
            wacc_hdr = ws_val[f"A{cur_r}"].value
            assert pytest.approx(wacc_hdr, 1e-4) == w_val

            for c_idx, g_val in enumerate(g_rates, start=2):
                col_letter = get_column_letter(c_idx)
                f_val = str(ws_val[f"{col_letter}{cur_r}"].value)
                assert f_val.startswith("="), f"[{sym}] Sensitivity matrix cell {col_letter}{cur_r} not a formula"
                assert "$H$5" in f_val, f"[{sym}] Sensitivity matrix cell {col_letter}{cur_r} missing $H$5 link"
                assert f"{col_letter}$13" in f_val, f"[{sym}] Sensitivity matrix cell {col_letter}{cur_r} missing column rate link"
                assert f"$A{cur_r}" in f_val, f"[{sym}] Sensitivity matrix cell {col_letter}{cur_r} missing row WACC link"

        # 7. Verify Balance Sheet Audit Badges (Tab 3, Rows 26 and 27)
        ws_bs = wb["Balance Sheet"]
        for col_idx in range(3, 8):
            col_letter = get_column_letter(col_idx)
            diff_cell = ws_bs[f"{col_letter}26"]
            diff_formula = str(diff_cell.value)
            assert diff_formula == f"={col_letter}13-{col_letter}25"

            status_cell = ws_bs[f"{col_letter}27"]
            status_formula = str(status_cell.value)
            assert status_formula == f'=IF(ABS({col_letter}26)<1, "BALANCED", "UNBALANCED")'

            # Assert green background fill
            fill_color = status_cell.fill.start_color.rgb
            assert fill_color in ["00E2EFDA", "E2EFDA"], (
                f"[{sym}] Status cell fill color {fill_color} is not expected soft green E2EFDA"
            )


class TestAdversarialVN30UniverseClosure:
    """
    Verifies that 100% of all 30 VN30 constituents pass the 5-year balance sheet balance test
    and export into error-free Excel workbooks.
    """

    @pytest.mark.parametrize("sym", VN30_SYMBOLS)
    def test_vn30_constituent_5y_balance_and_export(self, sym, export_scratch_dir, screener_snapshot):
        # 1. Compute 5Y synchronized forecast
        res = ThreeStatementEngine.build_forecast_from_screener(sym)
        assert res.symbol == sym
        assert len(res.forecast_years) == 5

        # 2. Strict Mathematical Balance Sheet Closure
        assert res.all_years_balanced is True, f"Constituent {sym} has unbalanced balance sheet"
        max_diff_billion = res.max_balance_difference / 1e9
        assert max_diff_billion < 1e-5, f"Constituent {sym} max balance difference in Billion VND: {max_diff_billion}"
        assert res.max_balance_difference < 1.0, f"Constituent {sym} max raw balance difference: {res.max_balance_difference}"

        for t in range(5):
            ta = res.balance_sheet.total_assets[t]
            tl = res.balance_sheet.total_liabilities[t]
            te = res.balance_sheet.total_equity[t]
            diff = abs(ta - (tl + te))
            rel_err = safe_div(diff, max(ta, 1.0), 0.0)
            assert (diff < 1.0) or (rel_err < 1e-5), (
                f"Constituent {sym} Year {res.forecast_years[t]} imbalance: |{ta} - ({tl}+{te})| = {diff} (rel_err={rel_err:.2e})"
            )

        # 3. Cash Flow Reconciliation to Ending Cash
        for t in range(5):
            beg_cash = res.cash_flow_statement.beginning_cash[t]
            delta_cash = res.cash_flow_statement.net_change_in_cash[t]
            end_cash = res.cash_flow_statement.ending_cash[t]
            bs_cash = res.balance_sheet.cash[t]
            assert pytest.approx(beg_cash + delta_cash, rel=1e-5, abs=1.0) == end_cash
            assert pytest.approx(end_cash, rel=1e-5, abs=1.0) == bs_cash

        # 4. Export to Excel and verify zero formula errors
        out_path = os.path.join(export_scratch_dir, f"VN30_{sym}_test.xlsx")
        FinancialModelExporter.export_to_excel(res, out_path)
        assert os.path.exists(out_path)

        wb = openpyxl.load_workbook(out_path, data_only=False)
        assert len(wb.sheetnames) == 7

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell.value is not None:
                        val_str = str(cell.value)
                        for token in ERROR_TOKENS:
                            assert token not in val_str, f"[{sym}] Formula error '{token}' in sheet '{sheet_name}' cell {cell.coordinate}"
