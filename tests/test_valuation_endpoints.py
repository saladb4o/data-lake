"""
Integration Tests for Server Endpoints:
- GET /api/valuation/comprehensive/{symbol}
- GET /api/backtest/fair_value/presets
- GET/POST /api/backtest/fair_value/run
"""

import pytest
from fastapi.testclient import TestClient
from server import app


@pytest.fixture
def client():
    return TestClient(app)


def test_api_get_comprehensive_valuation(client):
    """Test comprehensive valuation API for HPG."""
    resp = client.get("/api/valuation/comprehensive/HPG")
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "success"
    val = data["data"]
    assert val["symbol"] == "HPG"
    assert val["composite_fair_value"] > 0
    assert len(val["models"]) == 22
    assert "wacc_result" in val
    assert "risk_firewall" in val
    assert "scenarios" in val
    assert len(val["scenarios"]["sensitivity_grid_5x5"]) == 5


def test_api_get_fair_value_backtest_presets(client):
    """Test retrieval of backtest presets and models."""
    resp = client.get("/api/backtest/fair_value/presets")
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "success"
    presets = data["data"]
    assert len(presets["modes"]) == 3
    assert len(presets["screening_presets"]) >= 5
    assert len(presets["valuation_models"]) >= 10


def test_api_run_fair_value_backtest(client):
    """Test running 3-Mode Backtest via API endpoint."""
    payload = {
        "mode": "hybrid_funnel",
        "screening_strategy": "peter_lynch_garp",
        "valuation_model_id": "composite_fair_value",
        "margin_of_safety_pct": 15.0,
        "exit_premium_pct": 20.0,
        "start_year": 2023,
        "end_year": 2025,
    }
    resp = client.post("/api/backtest/fair_value/run", params=payload)
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "success"
    res = data["data"]
    assert res["mode"] == "hybrid_funnel"
    assert "metrics" in res
    assert "yearly_returns" in res
    assert "equity_curve" in res
    assert "model_tournament_matrix" in res


# =============================================================================
# 3-WAY INTEGRATED FORECAST & EXCEL EXPORT REST ENDPOINTS
# =============================================================================

def test_api_get_three_statement_forecast_hpg(client):
    """Test GET /api/valuation/3-way-forecast/HPG returns valid 5Y integrated model."""
    resp = client.get("/api/valuation/3-way-forecast/HPG")
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "success"
    fc = data["data"]
    assert fc["symbol"] == "HPG"
    assert fc["start_year"] == 2026
    assert fc["forecast_years"] == [2026, 2027, 2028, 2029, 2030]
    assert fc["all_years_balanced"] is True
    assert fc["max_balance_difference"] < 1.0

    # Income Statement
    is_stmt = fc["income_statement"]
    assert len(is_stmt["revenue"]) == 5
    assert len(is_stmt["gross_profit"]) == 5
    assert len(is_stmt["ebit"]) == 5
    assert len(is_stmt["npat"]) == 5

    # Balance Sheet
    bs = fc["balance_sheet"]
    assert len(bs["total_assets"]) == 5
    assert len(bs["total_liabilities"]) == 5
    assert len(bs["total_equity"]) == 5
    assert bs["is_balanced"] == [True, True, True, True, True]

    # Cash Flow Statement
    cfs = fc["cash_flow_statement"]
    assert len(cfs["net_cfo"]) == 5
    assert len(cfs["net_cfi"]) == 5
    assert len(cfs["net_cff"]) == 5
    assert len(cfs["ending_cash"]) == 5

    # Schedules & Distress
    assert len(fc["working_capital_schedule"]) == 5
    assert len(fc["debt_schedule"]) == 5
    assert "liquidity_distress_check" in fc
    assert fc["liquidity_distress_check"]["summary_assessment"] in ("HEALTHY", "TIGHT", "DISTRESSED")


def test_api_get_three_statement_forecast_with_parameters(client):
    """Test GET /api/valuation/3-way-forecast/{symbol} with query parameters."""
    resp = client.get("/api/valuation/3-way-forecast/FPT?start_year=2027&tax_rate=0.25")
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "success"
    fc = data["data"]
    assert fc["symbol"] == "FPT"
    assert fc["start_year"] == 2027
    assert fc["forecast_years"] == [2027, 2028, 2029, 2030, 2031]
    assert fc["all_years_balanced"] is True


def test_api_get_three_statement_forecast_financial_sector(client):
    """Test GET /api/valuation/3-way-forecast/VCB for financial sector bank isolation."""
    resp = client.get("/api/valuation/3-way-forecast/VCB")
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "success"
    fc = data["data"]
    assert fc["is_financial_sector"] is True
    assert fc["all_years_balanced"] is True


def test_api_export_excel_endpoint_hpg(client):
    """Test GET /api/valuation/export-excel/HPG streams a valid 7-tab openpyxl workbook."""
    import io
    import openpyxl

    resp = client.get("/api/valuation/export-excel/HPG")
    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers.get("content-type", "")
    assert "attachment; filename=" in resp.headers.get("content-disposition", "")
    assert "HPG_3Way_Financial_Model.xlsx" in resp.headers.get("content-disposition", "")

    # Parse streaming content into openpyxl workbook
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=False)
    expected_sheets = [
        "Summary & Dashboard",
        "Income Statement",
        "Balance Sheet",
        "Cash Flow Statement",
        "Working Capital Schedule",
        "Debt & Capital Schedule",
        "Valuation & Sensitivity",
    ]
    assert len(wb.sheetnames) == 7
    for name in expected_sheets:
        assert name in wb.sheetnames


def test_api_export_excel_endpoint_raw_scale_fpt(client):
    """Test GET /api/valuation/export-excel/FPT with scale_unit=raw."""
    import io
    import openpyxl

    resp = client.get("/api/valuation/export-excel/FPT?scale_unit=raw&start_year=2026")
    assert resp.status_code == 200
    assert len(resp.content) > 5000

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=False)
    assert len(wb.sheetnames) == 7
    assert wb.active.title == "Summary & Dashboard"



def test_error_response_maps_value_error_to_422():
    """A ValueError is a data gap the caller can act on, not a server fault."""
    from server import _error_response

    assert _error_response(ValueError("no price")).status_code == 422
    assert _error_response(RuntimeError("boom")).status_code == 500
    assert _error_response(KeyError("k")).status_code == 500
