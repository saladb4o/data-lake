"""
=============================================================================
MODANO 3-WAY INTEGRATED MODELING ECOSYSTEM: EXCEL EXPORTER TEST SUITE
=============================================================================
Comprehensive test suite validating the Modano-compliant 7-Tab Excel Exporter:
1. File generation, workbook initialization, and binary integrity
2. 7-Tab Sheet Architecture & Tab Layout Compliance
3. Dynamic Native Excel Formulas (SUM, IF, cross-sheet references, IS_BALANCED)
4. Corporate Finance Visual Formatting (Navy Blue headers, cell borders, number formats)
5. Multi-Sector VN30 Constituent Export (FPT, HPG, VCB, MWG, VIC)
=============================================================================
"""

import os
import pytest
import openpyxl
from openpyxl.styles import PatternFill

from services.three_statement_engine import ThreeStatementEngine
from services.financial_model_exporter import FinancialModelExporter, COLOR_NAVY_HEADER
from services.stock_service import VN30_SYMBOLS


@pytest.fixture
def temp_export_dir(tmp_path):
    """Provides a temporary directory for test spreadsheet artifacts."""
    export_dir = tmp_path / "model_exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    return str(export_dir)


@pytest.fixture
def fpt_forecast_result():
    """Builds a verified 5-year forecast for FPT."""
    return ThreeStatementEngine.build_forecast_from_screener("FPT")


@pytest.fixture
def hpg_forecast_result():
    """Builds a verified 5-year forecast for HPG."""
    return ThreeStatementEngine.build_forecast_from_screener("HPG")


@pytest.fixture
def vcb_forecast_result():
    """Builds a verified 5-year forecast for VCB (Financial sector)."""
    return ThreeStatementEngine.build_forecast_from_screener("VCB")


# =============================================================================
# TIER 1: WORKBOOK GENERATION & FILE INTEGRITY
# =============================================================================

class TestTier1WorkbookGeneration:
    """Tier 1: Workbook creation, path resolution, and non-empty file validation."""

    def test_export_generates_valid_file(self, fpt_forecast_result, temp_export_dir):
        out_file = os.path.join(temp_export_dir, "FPT_3way_model.xlsx")
        saved_path = FinancialModelExporter.export_to_excel(fpt_forecast_result, out_file)
        
        assert os.path.exists(saved_path)
        assert os.path.getsize(saved_path) > 5000 # Non-trivial binary size (>5KB)
        assert saved_path.endswith(".xlsx")

    def test_raw_unit_scale_export(self, hpg_forecast_result, temp_export_dir):
        out_file = os.path.join(temp_export_dir, "HPG_raw_model.xlsx")
        saved_path = FinancialModelExporter.export_to_excel(hpg_forecast_result, out_file, scale_unit="raw")
        assert os.path.exists(saved_path)
        assert os.path.getsize(saved_path) > 5000


# =============================================================================
# TIER 2: 7-TAB STRUCTURE & SHEET ARCHITECTURE
# =============================================================================

class TestTier2SheetArchitecture:
    """Tier 2: Verification of all 7 Modano-compliant sheets."""

    EXPECTED_SHEETS = [
        "Summary & Dashboard",
        "Income Statement",
        "Balance Sheet",
        "Cash Flow Statement",
        "Working Capital Schedule",
        "Debt & Capital Schedule",
        "Valuation & Sensitivity",
    ]

    def test_exact_7_tab_architecture(self, fpt_forecast_result, temp_export_dir):
        out_file = os.path.join(temp_export_dir, "FPT_tabs_test.xlsx")
        FinancialModelExporter.export_to_excel(fpt_forecast_result, out_file)
        
        wb = openpyxl.load_workbook(out_file, data_only=False)
        assert len(wb.sheetnames) == 7
        for expected_name in self.EXPECTED_SHEETS:
            assert expected_name in wb.sheetnames, f"Missing sheet: {expected_name}"

    def test_dashboard_is_active_sheet(self, fpt_forecast_result, temp_export_dir):
        out_file = os.path.join(temp_export_dir, "FPT_active_tab_test.xlsx")
        FinancialModelExporter.export_to_excel(fpt_forecast_result, out_file)
        
        wb = openpyxl.load_workbook(out_file, data_only=False)
        assert wb.active.title == "Summary & Dashboard"


# =============================================================================
# TIER 3: DYNAMIC FORMULAS & CROSS-SHEET INTEGRITY
# =============================================================================

class TestTier3DynamicFormulas:
    """Tier 3: Verification of live Excel formulas and cross-sheet links."""

    def test_income_statement_formulas(self, fpt_forecast_result, temp_export_dir):
        out_file = os.path.join(temp_export_dir, "FPT_is_formulas.xlsx")
        FinancialModelExporter.export_to_excel(fpt_forecast_result, out_file)
        
        wb = openpyxl.load_workbook(out_file, data_only=False)
        ws_is = wb["Income Statement"]
        
        # Check Gross Profit formula: =C5-C7
        gp_formula = str(ws_is["C8"].value)
        assert gp_formula.startswith("=")
        assert "C5" in gp_formula and "C7" in gp_formula

        # Check EBIT formula: =C11-C12
        ebit_formula = str(ws_is["C13"].value)
        assert ebit_formula.startswith("=")
        assert "C11" in ebit_formula and "C12" in ebit_formula

        # Check NPAT formula: =C17-C18
        npat_formula = str(ws_is["C20"].value)
        assert npat_formula.startswith("=")
        assert "C17" in npat_formula and "C18" in npat_formula

    def test_balance_sheet_closure_formulas_and_checks(self, hpg_forecast_result, temp_export_dir):
        out_file = os.path.join(temp_export_dir, "HPG_bs_formulas.xlsx")
        FinancialModelExporter.export_to_excel(hpg_forecast_result, out_file)
        
        wb = openpyxl.load_workbook(out_file, data_only=False)
        ws_bs = wb["Balance Sheet"]
        
        # Check Total Current Assets: =SUM(C5:C8)
        tca_formula = str(ws_bs["C9"].value)
        assert "SUM(" in tca_formula

        # Check Balance Check Difference formula
        diff_formula = str(ws_bs["C26"].value)
        assert diff_formula.startswith("=")
        assert "C13" in diff_formula and "C25" in diff_formula

        # Check IS_BALANCED IF condition
        status_formula = str(ws_bs["C27"].value)
        assert "IF(" in status_formula
        assert "BALANCED" in status_formula

    def test_cash_flow_cross_sheet_links(self, fpt_forecast_result, temp_export_dir):
        out_file = os.path.join(temp_export_dir, "FPT_cfs_links.xlsx")
        FinancialModelExporter.export_to_excel(fpt_forecast_result, out_file)
        
        wb = openpyxl.load_workbook(out_file, data_only=False)
        ws_cfs = wb["Cash Flow Statement"]
        
        # Check cross-sheet link to Income Statement and Working Capital
        cust_formula = str(ws_cfs["C6"].value)
        assert "Income Statement" in cust_formula
        assert "Working Capital Schedule" in cust_formula

        # Check Ending Cash formula: =C25+C26
        end_cash_formula = str(ws_cfs["C27"].value)
        assert end_cash_formula.startswith("=")
        assert "C25" in end_cash_formula and "C26" in end_cash_formula

    def test_valuation_sensitivity_2d_matrix_formulas(self, fpt_forecast_result, temp_export_dir):
        out_file = os.path.join(temp_export_dir, "FPT_sensitivity_test.xlsx")
        FinancialModelExporter.export_to_excel(fpt_forecast_result, out_file)
        
        wb = openpyxl.load_workbook(out_file, data_only=False)
        ws_val = wb["Valuation & Sensitivity"]
        
        # Check dynamic cell formula in the 5x5 sensitivity matrix
        matrix_cell_val = str(ws_val["D15"].value)
        assert matrix_cell_val.startswith("=")
        assert "$H$5" in matrix_cell_val # Reference to base valuation numerator


# =============================================================================
# TIER 4: CORPORATE FINANCE VISUAL STYLING & NUMBER FORMATS
# =============================================================================

class TestTier4FormattingAndStyling:
    """Tier 4: Verification of font colors, navy headers, and number formatting."""

    def test_header_navy_styling(self, fpt_forecast_result, temp_export_dir):
        out_file = os.path.join(temp_export_dir, "FPT_styles_test.xlsx")
        FinancialModelExporter.export_to_excel(fpt_forecast_result, out_file)
        
        wb = openpyxl.load_workbook(out_file, data_only=False)
        ws_dash = wb["Summary & Dashboard"]
        
        # Check Title Banner Font and Fill
        assert ws_dash["A1"].font.bold is True
        assert ws_dash["A1"].fill.start_color.rgb == f"00{COLOR_NAVY_HEADER}" or ws_dash["A1"].fill.start_color.rgb == COLOR_NAVY_HEADER

    def test_number_formats_applied(self, hpg_forecast_result, temp_export_dir):
        out_file = os.path.join(temp_export_dir, "HPG_num_fmt_test.xlsx")
        FinancialModelExporter.export_to_excel(hpg_forecast_result, out_file)
        
        wb = openpyxl.load_workbook(out_file, data_only=False)
        ws_is = wb["Income Statement"]
        
        # Revenue should have Billion / integer format
        rev_cell = ws_is["C5"]
        assert rev_cell.number_format is not None
        assert "#,##0" in rev_cell.number_format

        # Revenue Growth should have % format
        rev_g_cell = ws_is["C6"]
        assert "%" in rev_g_cell.number_format


# =============================================================================
# TIER 5: MULTI-SECTOR VN30 CONSTITUENT SWEEP
# =============================================================================

class TestTier5VN30ExportSweep:
    """Tier 5: Validates that multiple diverse VN30 constituents export cleanly."""

    @pytest.mark.parametrize("sym", ["FPT", "HPG", "VCB", "MWG", "VIC"])
    def test_vn30_sample_constituents_export_successfully(self, sym, temp_export_dir):
        res = ThreeStatementEngine.build_forecast_from_screener(sym)
        out_file = os.path.join(temp_export_dir, f"{sym}_model.xlsx")
        saved_path = FinancialModelExporter.export_to_excel(res, out_file)
        
        assert os.path.exists(saved_path)
        assert os.path.getsize(saved_path) > 5000
        
        wb = openpyxl.load_workbook(saved_path, data_only=False)
        assert len(wb.sheetnames) == 7

    @pytest.mark.parametrize("sym", ["HPG", "FPT", "MWG", "VCB"])
    def test_zero_formula_errors_across_all_sheets(self, sym, temp_export_dir):
        """Validates that zero formula errors (#REF!, #NAME?, #VALUE!, #DIV/0!) exist across all 7 sheets."""
        res = ThreeStatementEngine.build_forecast_from_screener(sym)
        out_file = os.path.join(temp_export_dir, f"{sym}_formula_audit.xlsx")
        saved_path = FinancialModelExporter.export_to_excel(res, out_file)
        
        wb = openpyxl.load_workbook(saved_path, data_only=False)
        error_tokens = ["#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"]
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    val_str = str(cell.value) if cell.value is not None else ""
                    for token in error_tokens:
                        assert token not in val_str, f"Found formula error '{token}' in sheet '{sheet_name}' cell {cell.coordinate}: {val_str}"

